from tkinter import *
from tkinter import ttk
import urllib.request as requests
from bs4 import BeautifulSoup
import pandas as pd
import global_settings as gs
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from src.components.settings import system

class Receipt:
    # Class attribute
    receipt_data = {
                "found": 404,
                "place name": "null",
                "place cnpj": "null",
                "place address": "null",
                "total items": "null",
                "total value": "null",
                "discount": "null",
                "final value": "null",
                "date": "null",
                "buyer": "null",
                "payment": "null",
                "receipt key": "null"
    }
    item_data = [{
        "category": "null",
        "description": "null",
        "type": "null",
        "qtd": 0,
        "unit": "null",
        "unit cost": 0.0,
        "total cost": 0.0
    }]
    receipt_header = (
        "Date", "Place Name", "Category", "Description", "Type",
        "Qtd", "Unit", "Unit Cost", "Total Cost", "Receipt Key"
    )

    # Class methods

    #region MANIPULATE RECEIPT DATA
    def getDataStatus(self):
        return self.receipt_data['found']

    def printReceiptData(self, frame: Frame):
        key_receipt_list = list(self.receipt_data.keys())
        val_receipt_list = list(self.receipt_data.values())
        
        for i in range(1, len(self.receipt_data)):
            # Organize Receipt Items
            if (key_receipt_list[i] == "total items"):
                key_item_list = list(self.item_data[0].keys())
                itemsTable =  ttk.Treeview(frame, column=key_item_list, show='headings', height=15)
                
                # Add Header Name Reference
                for k in range(len(key_item_list)):
                    if key_item_list[k] == 'description':
                        itemsTable.column(key_item_list[k], anchor=CENTER, width=350)
                    elif key_item_list[k] == 'qtd' or key_item_list[k] == 'unit':
                        itemsTable.column(key_item_list[k], anchor=CENTER, width=50)
                    else:
                        itemsTable.column(key_item_list[k], anchor=CENTER, width=100)

                # Add Header
                for k in range(len(key_item_list)):
                    itemsTable.heading(key_item_list[k], text=str(key_item_list[k]), anchor=CENTER)

                for j in range(len(self.item_data)):
                    val_item_list = list(self.item_data[j].values())
                    itemsTable.insert(parent='', index=j, iid=j, text='', values=val_item_list)
                
                itemsTable.pack()

            elif i == 1:
                # Table for place information

                placeTable =  ttk.Treeview(frame, column=('0', '1'), show="tree", height=3)                
                
                placeTable.column('0', width=80)
                placeTable.column('1', width=600)
                
                # Add content
                for k in range(1, 4):
                    value = (key_receipt_list[k], val_receipt_list[k])
                    placeTable.insert(parent='', index='end', iid=k, text='', values=value)

                placeTable.pack()

            elif i > 4:
                # Table for receipt details

                detailTable =  ttk.Treeview(frame, column=('0', '1'), show="tree", height=7)
                detailTable.column('0', width=80)
                detailTable.column('1', width=600)

                for k in range(i, len(val_receipt_list)):
                    value = (key_receipt_list[k], val_receipt_list[k])
                    detailTable.insert(parent='', index='end', iid=k, text='', values=value)
                    detailTable.pack()
                
                return
            else:
                continue

        return
    
    def getReceiptData(self, readAddress: str):
        # Autor: View || Data: 2022/12/15
        def fixSpaces(address: str):
            address = address.strip()
            flag = FALSE
            newAddress = ''
            for char in address:
                if char == '\t' or char == '\n':
                    continue
                elif char == ' ':
                    if flag == TRUE:
                        newAddress += char
                        flag = FALSE
                elif char == ',':
                    newAddress += ', '
                else:
                    newAddress += char
                    flag = TRUE

            return newAddress.replace(', , ', ', ')

        if(readAddress):
            with requests.urlopen(readAddress) as res:
                body = res.read()
                soup = BeautifulSoup(body, 'html.parser')

                # Get Place informations
                receipt_title = soup.find(id="conteudo").select('div')[1]
                place_name = receipt_title.select('div')[0].text.strip()
                place_cnpj = receipt_title.select('div')[1].text
                place_cnpj = place_cnpj.replace('CNPJ:', '').strip()

                place_address = receipt_title.select('div')[2].text.strip()
                place_address = fixSpaces(place_address)

                # Get bougth items informations
                all_tr_data = []
                for item_row in soup.find(id="tabResult").select('tr'):
                    item_name = item_row.select('td')[0].select('span')[0].text.strip()

                    item_qtd = item_row.select('td')[0].select('span')[2].text.strip()
                    item_qtd = item_qtd.replace('Qtde.:', '')
                    item_qtd = float(item_qtd.replace(',','.'))


                    item_unit = item_row.select('td')[0].select('span')[3].text.strip()
                    item_unit = item_unit.replace('UN: ', '')

                    item_unit_price = item_row.select('td')[0].select('span')[4].text.strip()
                    item_unit_price = item_unit_price.replace('Vl. Unit.:', '')
                    item_unit_price = float(item_unit_price.replace(',','.'))
                    
                    item_total = item_row.select('td')[1].select('span')[0].text.strip()
                    item_total = float(item_total.replace(',','.'))

                    all_tr_data.append({
                        "category": "null",
                        "description": item_name,
                        "type": "null",
                        "qtd": item_qtd,
                        "unit": item_unit,
                        "unit cost": item_unit_price, 
                        "total cost": item_total
                    })

                # receipt total
                receipt_total = soup.find(id="totalNota").select('div')

                if(len(receipt_total) > 5):
                    total_value = receipt_total[1].select('div > span')[0].text.strip()
                    total_value = float(total_value.replace(',','.'))

                    discount = receipt_total[2].select('div > span')[0].text.strip()
                    discount = float(discount.replace(',','.'))
                    
                    final_value = receipt_total[3].select('div > span')[0].text.strip()
                    final_value = float(final_value.replace(',','.'))
                
                else:
                    discount = 0

                    final_value = receipt_total[1].select('div > span')[0].text.strip()
                    final_value = float(final_value.replace(',','.'))
                    total_value = final_value

                total_items = receipt_total[0].select('div > span')[0].text.strip()
                total_items = float(total_items.replace(',','.'))


                # Receipt Date
                receipt_date_info = soup.find(id="infos").select('div')[0].select('div > ul > li')[0].text.strip()
                date_end_index = receipt_date_info.find("  - Via Consumidor")
                receipt_date = ''
                for i in range((date_end_index - 19), date_end_index):
                    receipt_date += receipt_date_info[i]



                # Payment form
                if len(receipt_total) > 5:
                    payment_form = soup.find(id="totalNota").select('div')[5].select('div > label')[0].text.strip()

                else:
                    payment_form = soup.find(id="totalNota").select('div')[3].select('div > label')[0].text.strip()

                # Buyer Information
                if(len(soup.find(id="infos").select('div')[2].select('div > ul > li')) > 2):
                    receipt_buyer = soup.find(id="infos").select('div')[2].select('div > ul > li')[1].text.strip()
                    receipt_buyer = receipt_buyer.replace('Nome: ', '')

                else:
                    receipt_buyer = ""

                # Receipt Key
                receipt_key = soup.find(id="infos").select('div')[1].select('div > ul > li > span')[0].text.strip()

            self.receipt_data = {
                "found": 200,
                "place name": place_name,
                "place cnpj": place_cnpj,
                "place address": place_address,
                "total items": total_items,
                "total value": total_value,
                "discount": discount,
                "final value": final_value,
                "date": receipt_date,
                "buyer": receipt_buyer,
                "payment": payment_form,
                "receipt key": receipt_key
            }
            self.item_data = all_tr_data

        else:
            print("QR Code address cannot exist")
            return FALSE

        return TRUE
    
    def entryBoxes(self, frame: Frame):
        # TABLE PLACE
        placeFrame = Frame(frame)
        placeFrame.pack()

        placename_label = Label(placeFrame, text="place name")
        placename_label.grid(row=0, column=0)
        placename_entry = Entry(placeFrame, width=60)
        placename_entry.grid(row=0, column=1)
        placename_entry.insert(0, self.receipt_data['place name'])

        placecnpj_label = Label(placeFrame, text="place cnpj")
        placecnpj_label.grid(row=1, column=0)
        placecnpj_entry = Entry(placeFrame, width=60)
        placecnpj_entry.grid(row=1, column=1)
        placecnpj_entry.insert(0, self.receipt_data['place cnpj'])

        placeaddress_label = Label(placeFrame, text="place address")
        placeaddress_label.grid(row=2, column=0)
        placeaddress_entry = Entry(placeFrame, width=60)
        placeaddress_entry.grid(row=2, column=1)
        placeaddress_entry.insert(0, self.receipt_data['place address'])


        # TABLE RECEIPT ITEM
        itemFrame = Frame(frame)
        itemFrame.pack()
                
        category_label = Label(itemFrame, text="category", height=2, width=10)
        category_label.grid(row=0, column=0)
        description_label = Label(itemFrame, text="description", height=2, width=35)
        description_label.grid(row=0, column=1)
        type_label = Label(itemFrame, text="type", height=2)
        type_label.grid(row=0, column=2)
        qtd_label = Label(itemFrame, text="qtd", height=2, width=2)
        qtd_label.grid(row=0, column=3)
        unit_label = Label(itemFrame, text="unit", height=2, width=2)
        unit_label.grid(row=0, column=4)
        unitcost_label = Label(itemFrame, text="unit cost", height=2)
        unitcost_label.grid(row=0, column=5)
        totalcost_label = Label(itemFrame, text="total cost", height=2)
        totalcost_label.grid(row=0, column=6)

        for i, row in enumerate(self.item_data):
            j = i + 1
            category_entry = Entry(itemFrame)
            category_entry.grid(row=j, column=0)
            category_entry.insert(i, row['category'])
            description_entry = Entry(itemFrame, width=60)
            description_entry.grid(row=j, column=1)
            description_entry.insert(i, row['description'])
            type_entry = Entry(itemFrame)
            type_entry.grid(row=j, column=2)
            type_entry.insert(i, row['type'])
            qtd_entry = Entry(itemFrame, width=6)
            qtd_entry.grid(row=j, column=3)
            qtd_entry.insert(i, row['qtd'])
            unit_entry = Entry(itemFrame, width=6)
            unit_entry.grid(row=j, column=4)
            unit_entry.insert(i, row['unit'])
            unitcost_entry = Entry(itemFrame)
            unitcost_entry.grid(row=j, column=5)
            unitcost_entry.insert(i, row['unit cost'])
            totalcost_entry = Entry(itemFrame)
            totalcost_entry.grid(row=j, column=6)
            totalcost_entry.insert(i, row['total cost'])

        # TABLE RECEIPT DETAILS
        receiptFrame = Frame(frame)
        receiptFrame.pack()

        totalvalue_label = Label(receiptFrame, text="total value")
        totalvalue_label.grid(row=0, column=0)
        totalvalue_entry = Entry(receiptFrame, width=60)
        totalvalue_entry.grid(row=0, column=1)
        totalvalue_entry.insert(0, self.receipt_data['total value'])
        
        discount_label = Label(receiptFrame, text="discount")
        discount_label.grid(row=1, column=0)
        discount_entry = Entry(receiptFrame, width=60)
        discount_entry.grid(row=1, column=1)
        discount_entry.insert(0, self.receipt_data['discount'])
        
        finalvalue_label = Label(receiptFrame, text="final value")
        finalvalue_label.grid(row=2, column=0)
        finalvalue_entry = Entry(receiptFrame, width=60)
        finalvalue_entry.grid(row=2, column=1)
        finalvalue_entry.insert(0, self.receipt_data['final value'])
        
        date_label = Label(receiptFrame, text="date")
        date_label.grid(row=3, column=0)
        date_entry = Entry(receiptFrame, width=60)
        date_entry.grid(row=3, column=1)
        date_entry.insert(0, self.receipt_data['date'])
        
        buyer_label = Label(receiptFrame, text="buyer")
        buyer_label.grid(row=4, column=0)
        buyer_entry = Entry(receiptFrame, width=60)
        buyer_entry.grid(row=4, column=1)
        buyer_entry.insert(0, self.receipt_data['buyer'])
        
        payment_label = Label(receiptFrame, text="payment")
        payment_label.grid(row=5, column=0)
        payment_entry = Entry(receiptFrame, width=60)
        payment_entry.grid(row=5, column=1)
        payment_entry.insert(0, self.receipt_data['payment'])
        
        receiptkey_label = Label(receiptFrame, text="receipt key")
        receiptkey_label.grid(row=6, column=0)
        receiptkey_entry = Entry(receiptFrame, width=60)
        receiptkey_entry.grid(row=6, column=1)
        receiptkey_entry.insert(0, self.receipt_data['receipt key'])

    def saveTable(self, frame: Frame):
        table = 0
        for widget in frame.winfo_children():
            itemTable = 0
            # Check if object is a frame
            if isinstance(widget, Frame):
                labelName = ''                
                itemRow = 0
                for widgetFrame in widget.winfo_children():

                    # Check if object is an Entry
                    if isinstance(widgetFrame, Entry):
                        if table == 0 or table == 2:
                            self.receipt_data[labelName] = widgetFrame.get()
                        if table == 1:
                            getHeader = list(self.item_data[0].keys())

                            self.item_data[itemTable][getHeader[itemRow]] = widgetFrame.get()

                            # Check if row end
                            if itemRow > 5:
                                itemRow =  0 
                                itemTable += 1
                            else :
                                itemRow += 1
                            

                    elif isinstance (widgetFrame, Label):
                        labelName = widgetFrame.cget("text")

                table += 1

    #endregion

    #region MANIPULATE SHEETS
    def createReceiptHistory(self):
        if (system.checkDirectoryExist(gs.__expenses_path__)):
            # creating a blank Workbook
            book = openpyxl.Workbook()
            sheet = book.active

            # add header
            header = (
                'Date', 'Place Name', 'Place CNPJ', 'Place Address',
                'Total Items', 'Total Value', 'Discount', 'Final Value',
                'Buyer', 'Payment', 'Receipt Key'
            )
            sheet.append(header)

            # Creating a table
            tab = Table(displayName='receipt_history', ref="A1:K2")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            sheet.add_table(tab)

            # Saving the file
            book.save(gs.__receipt_history__)

    def createUpdateFile(self):
        if system.checkDirectoryExist(gs.__expenses_path__):
            #creating a blank Workbook
            book = openpyxl.Workbook()
            sheet = book.create_sheet(gs.__expenses_page__)

            # add header
            header = (
                'Month', 'Date', 'Place Name',
                'Category', 'Description', 'Type',
                'Qtd', 'Unit', 'Unit Cost', 'Total Cost',
                'Receipt Key'
                )
            sheet.append(header)

            # Creating a table
            tab = Table(displayName='month_expenses', ref="A1:K2")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tab.tableStyleInfo = style
            sheet.add_table(tab)

            # Saving the file
            book.save(gs.__expenses_update__)

    def updateSheets(self):
        def writerBackUp():
            #TODO check if user has the directory 
            #check if directory to store backup does not exist, if not, create
            if not (system.checkDirectoryExist(gs.__expenses_path__)):
                system.createDirectory(gs.__expenses_path__)
                
            if not (system.checkDirectoryExist(gs.__expend_backup__)):
                system.createDirectory(gs.__expend_backup__)
            
            #check if file to store receipt history does not exist, if not, create
            if not (system.checkFileExist(gs.__receipt_history__)):
                self.createReceiptHistory()
                
            
            receipt_date =  self.receipt_data["date"].replace('/', '_').replace(' ', '_').replace(':', '_')           
            filename = gs.__expend_backup__ + receipt_date + ".csv"

            data = []
            receipt = self.receipt_data

            for i in range(len(self.item_data)):
                item = self.item_data[i]

                item_row = {
                    'Date': receipt['date'],
                    'Place Name': receipt['place name'],                    
                    'Category': item['category'],
                    'Description': item['description'],
                    'Type': item['type'],
                    'Qtd': float(item['qtd']),
                    'Unit': item['unit'],
                    'Unit Cost': float(item['unit cost']),
                    'Total Cost': float(item['total cost']),
                    "Receipt Key": receipt['receipt key']
                }
                
                data.append(item_row)
            
            pd.DataFrame(data).to_csv(filename, columns=self.receipt_header, index=False)

            # Store receipt summary
            receipt_infos = {
                'Date': receipt['date'],
                'Place Name': receipt['place name'],
                'Place CNPJ': receipt['place cnpj'],
                'Place Address': receipt['place address'],
                'Total Items': int(receipt['total items']),
                'Total Value': float(receipt['total value']),
                'Discount': float(receipt['discount']),
                'Final Value': float(receipt['final value']),
                'Buyer': receipt['buyer'],
                'Payment': receipt['payment'],
                'Receipt Key': receipt['receipt key']
            }
            book = openpyxl.load_workbook(gs.__receipt_history__)
            sheet = book.active
            sheet.append(list(receipt_infos.values()))

            book.save(gs.__receipt_history__)

            return

        def getRowValues(data: dict):
            val = list(data.values())
            return val

        def updateSheet():
            if not system.checkFileExist(gs.__expenses_update__):
                self.createUpdateFile()

            book = openpyxl.load_workbook(gs.__expenses_update__)
            sheet = book[gs.__expenses_page__]

            data = []
            receipt = self.receipt_data

            for i in range(len(self.item_data)):
                item = self.item_data[i]
                item_row = {
                    'Month': '',
                    'Date': receipt['date'],
                    'Place Name': receipt['place name'],
                    'Category': item['category'],
                    'Description': item['description'],
                    'Type': item['type'],
                    'Qtd': float(item['qtd']),
                    'Unit': item['unit'],
                    'Unit Cost': float(item['unit cost']),
                    'Total Cost': float(item['total cost']),
                    "Receipt Key": receipt['receipt key']
                }

                data.append(item_row)

            for row in data:
                val = getRowValues(row)
                sheet.append(val)

            book.save(gs.__expenses_update__)
    
            return

        writerBackUp()
        updateSheet()

    #endregion